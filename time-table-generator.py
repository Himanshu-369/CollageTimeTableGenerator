import customtkinter as ctk
import sqlite3
import random
from tkinter import messagebox, filedialog
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

# --- CONFIGURATION & THEME ---
ctk.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
ctk.set_default_color_theme("blue")  # Themes: "blue", "green", "dark-blue"

# --- DATABASE MANAGER ---
class DatabaseManager:
    def __init__(self, db_name="timetable.db"):
        self.conn = sqlite3.connect(db_name)
        self.cursor = self.conn.cursor()
        self.create_tables()

    def create_tables(self):
        # Teachers Table
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS teachers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                code TEXT UNIQUE
            )
        """)
        # Subjects Table
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS subjects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                code TEXT,
                type TEXT, -- 'Theory' or 'Lab'
                hours_per_week INTEGER
            )
        """)
        # Rooms Table
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS rooms (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE,
                capacity INTEGER,
                type TEXT -- 'Lecture Hall' or 'Lab'
            )
        """)
        # Sections Table
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS sections (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE
            )
        """)
        self.conn.commit()

    def add_teacher(self, name, code):
        try:
            self.cursor.execute("INSERT INTO teachers (name, code) VALUES (?, ?)", (name, code))
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False

    def add_subject(self, name, code, sub_type, hours):
        self.cursor.execute("INSERT INTO subjects (name, code, type, hours_per_week) VALUES (?, ?, ?, ?)", 
                            (name, code, sub_type, hours))
        self.conn.commit()

    def add_room(self, name, capacity, room_type):
        try:
            self.cursor.execute("INSERT INTO rooms (name, capacity, type) VALUES (?, ?, ?)", (name, capacity, room_type))
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False
            
    def add_section(self, name):
        try:
            self.cursor.execute("INSERT INTO sections (name) VALUES (?)", (name,))
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False

    def fetch_all(self, table):
        self.cursor.execute(f"SELECT * FROM {table}")
        return self.cursor.fetchall()

    def delete_record(self, table, record_id):
        self.cursor.execute(f"DELETE FROM {table} WHERE id=?", (record_id,))
        self.conn.commit()

# --- ALGORITHM ENGINE ---
class Scheduler:
    def __init__(self, db):
        self.db = db
        self.days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
        self.slots = ["1 (1.10-2.00)", "2 (2.00-2.50)", "3 (2.50-3.40)", "4 (3.40-4.30)", "5 (4.30-5.15)", "6 (5.15-6.00)"]
        
    def generate(self):
        teachers = self.db.fetch_all("teachers")
        subjects = self.db.fetch_all("subjects")
        rooms = self.db.fetch_all("rooms")
        sections = self.db.fetch_all("sections")
        
        if not (teachers and subjects and rooms and sections):
            return None, "Missing Data: Please add Teachers, Subjects, Rooms, and Sections."

        # Schedule Structure: schedule[day][slot][section_id] = {subject, teacher, room}
        schedule = {day: {slot: {} for slot in self.slots} for day in self.days}
        
        # Tracking to prevent collisions
        teacher_busy = {day: {slot: [] for slot in self.slots} for day in self.days}
        room_busy = {day: {slot: [] for slot in self.slots} for day in self.days}
        
        # Simple randomized greedy algorithm with backtracking simulation
        for section in sections:
            sec_id, sec_name = section
            
            # For this demo, we assume every section takes every subject (In a real app, you'd map Subject->Section)
            # We shuffle subjects to vary the output
            section_subjects = list(subjects) 
            random.shuffle(section_subjects)

            for sub in section_subjects:
                sub_id, sub_name, sub_code, sub_type, hours = sub
                hours_assigned = 0
                
                # Try to assign hours
                attempts = 0
                while hours_assigned < hours and attempts < 100:
                    day = random.choice(self.days)
                    
                    # Logic for Labs (Try to find consecutive slots) - simplified for demo
                    needed_slots = 2 if sub_type == 'Lab' else 1
                    
                    start_slot_idx = random.randint(0, len(self.slots) - needed_slots)
                    
                    # Check availability for the required block
                    can_schedule = True
                    selected_slots = []
                    
                    # Randomly assign a teacher and room capable of this subject
                    # (In a real app, map Teacher->Subject)
                    assigned_teacher = random.choice(teachers)
                    # Filter rooms by type (Lab vs Lecture)
                    valid_rooms = [r for r in rooms if r[3] == ('Lab' if sub_type == 'Lab' else 'Lecture Hall')]
                    if not valid_rooms: valid_rooms = rooms # Fallback
                    assigned_room = random.choice(valid_rooms)

                    for i in range(needed_slots):
                        slot = self.slots[start_slot_idx + i]
                        
                        # Constraints Check
                        if sec_id in schedule[day][slot]: can_schedule = False # Section busy
                        if assigned_teacher[0] in teacher_busy[day][slot]: can_schedule = False # Teacher busy
                        if assigned_room[0] in room_busy[day][slot]: can_schedule = False # Room busy
                        selected_slots.append(slot)
                    
                    if can_schedule:
                        for slot in selected_slots:
                            schedule[day][slot][sec_id] = {
                                "subject": f"{sub_name} ({sub_type})",
                                "teacher": assigned_teacher[1],
                                "room": assigned_room[1]
                            }
                            teacher_busy[day][slot].append(assigned_teacher[0])
                            room_busy[day][slot].append(assigned_room[0])
                        
                        hours_assigned += needed_slots
                    else:
                        attempts += 1
                        
        return schedule, "Success"

    def export_to_excel(self, schedule, filename):
        wb = openpyxl.Workbook()
        wb.remove(wb.active) # Remove default sheet
        
        # Get all sections to create sheets
        sections = self.db.fetch_all("sections")
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for sec_id, sec_name in sections:
            ws = wb.create_sheet(title=sec_name)
            
            # Header Info
            ws['B2'] = "School of Computer Application, JECRC University"
            ws['B3'] = "Time Table - Generated"
            ws['B4'] = f"Section: {sec_name}"
            ws['B2'].font = Font(bold=True, size=14)
            
            # Table Headers
            ws.cell(row=6, column=1).value = "Day / Time"
            for col, slot in enumerate(self.slots, start=2):
                cell = ws.cell(row=6, column=col)
                cell.value = slot
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

            # Fill Data
            for row, day in enumerate(self.days, start=7):
                day_cell = ws.cell(row=row, column=1)
                day_cell.value = day
                day_cell.font = Font(bold=True)
                day_cell.border = thin_border
                
                for col, slot in enumerate(self.slots, start=2):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    if sec_id in schedule[day][slot]:
                        data = schedule[day][slot][sec_id]
                        cell.value = f"{data['subject']}\n{data['teacher']}\n{data['room']}"
                    else:
                        cell.value = "---"

        wb.save(filename)

# --- GUI APPLICATION ---
class TimeTableApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        self.db = DatabaseManager()
        self.scheduler = Scheduler(self.db)
        self.schedule_data = None

        self.title("JECRC Timetable Generator")
        self.geometry("1100x700")

        # Layout Configuration
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.setup_sidebar()
        self.setup_pages()
        self.show_frame("Dashboard")

    def setup_sidebar(self):
        self.sidebar_frame = ctk.CTkFrame(self, width=200, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(8, weight=1)

        self.logo_label = ctk.CTkLabel(self.sidebar_frame, text="Timetable Gen", font=ctk.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        buttons = [
            ("Dashboard", "Dashboard"),
            ("Faculty", "Teachers"),
            ("Subjects", "Subjects"),
            ("Rooms", "Rooms"),
            ("Sections", "Sections"),
            ("Generate", "Generate"),
            ("Settings", "Settings")
        ]

        self.nav_buttons = {}
        for i, (text, name) in enumerate(buttons):
            btn = ctk.CTkButton(self.sidebar_frame, text=text, command=lambda n=name: self.show_frame(n))
            btn.grid(row=i+1, column=0, padx=20, pady=10)
            self.nav_buttons[name] = btn

    def setup_pages(self):
        self.frames = {}
        
        # --- Dashboard Frame ---
        dash = ctk.CTkFrame(self)
        self.frames["Dashboard"] = dash
        ctk.CTkLabel(dash, text="Dashboard Overview", font=("Arial", 24)).pack(pady=20)
        self.stats_label = ctk.CTkLabel(dash, text="Welcome! Use the sidebar to manage data.", font=("Arial", 16))
        self.stats_label.pack(pady=10)
        
        # --- Teachers Frame ---
        self.frames["Teachers"] = self.create_crud_frame("Faculty Management", 
                                                         ["Name", "Code"], 
                                                         self.add_teacher_action, 
                                                         "teachers")

        # --- Subjects Frame ---
        self.frames["Subjects"] = self.create_subject_frame()

        # --- Rooms Frame ---
        self.frames["Rooms"] = self.create_room_frame()

        # --- Sections Frame ---
        self.frames["Sections"] = self.create_crud_frame("Section Management", 
                                                         ["Section Name (e.g., BCA I A)"], 
                                                         self.add_section_action, 
                                                         "sections")

        # --- Generate Frame ---
        gen = ctk.CTkFrame(self)
        self.frames["Generate"] = gen
        ctk.CTkLabel(gen, text="Generate Timetable", font=("Arial", 24)).pack(pady=20)
        
        self.gen_btn = ctk.CTkButton(gen, text="Run Algorithm", command=self.run_generation, height=50, fg_color="green")
        self.gen_btn.pack(pady=20)

        self.preview_area = ctk.CTkTextbox(gen, width=800, height=400)
        self.preview_area.pack(pady=10)
        
        self.export_btn = ctk.CTkButton(gen, text="Export to Excel (.xlsx)", command=self.export_file, state="disabled")
        self.export_btn.pack(pady=10)

        # --- Settings Frame ---
        self.frames["Settings"] = self.create_settings_frame()

    def create_crud_frame(self, title, fields, add_command, table_name):
        frame = ctk.CTkFrame(self)
        ctk.CTkLabel(frame, text=title, font=("Arial", 24)).pack(pady=20)
        
        input_frame = ctk.CTkFrame(frame)
        input_frame.pack(pady=10)
        
        entries = []
        for field in fields:
            ent = ctk.CTkEntry(input_frame, placeholder_text=field, width=200)
            ent.pack(side="left", padx=5)
            entries.append(ent)
            
        ctk.CTkButton(input_frame, text="Add", command=lambda: add_command(entries, list_box)).pack(side="left", padx=5)
        
        list_box = ctk.CTkTextbox(frame, width=600, height=400)
        list_box.pack(pady=20)
        
        # Initial Load
        self.refresh_list(list_box, table_name)
        
        return frame

    def create_subject_frame(self):
        frame = ctk.CTkFrame(self)
        ctk.CTkLabel(frame, text="Subject Management", font=("Arial", 24)).pack(pady=20)
        
        input_frame = ctk.CTkFrame(frame)
        input_frame.pack(pady=10)
        
        name_ent = ctk.CTkEntry(input_frame, placeholder_text="Name")
        name_ent.pack(side="left", padx=5)
        code_ent = ctk.CTkEntry(input_frame, placeholder_text="Code")
        code_ent.pack(side="left", padx=5)
        type_ent = ctk.CTkComboBox(input_frame, values=["Theory", "Lab"])
        type_ent.pack(side="left", padx=5)
        hours_ent = ctk.CTkEntry(input_frame, placeholder_text="Hrs/Week")
        hours_ent.pack(side="left", padx=5)

        def add_sub():
            if name_ent.get() and hours_ent.get().isdigit():
                self.db.add_subject(name_ent.get(), code_ent.get(), type_ent.get(), int(hours_ent.get()))
                self.refresh_list(list_box, "subjects")
                name_ent.delete(0, 'end')
                code_ent.delete(0, 'end')
                hours_ent.delete(0, 'end')
            else:
                messagebox.showerror("Error", "Invalid Input")

        ctk.CTkButton(input_frame, text="Add", command=add_sub).pack(side="left", padx=5)
        list_box = ctk.CTkTextbox(frame, width=700, height=400)
        list_box.pack(pady=20)
        self.refresh_list(list_box, "subjects")
        return frame

    def create_room_frame(self):
        frame = ctk.CTkFrame(self)
        ctk.CTkLabel(frame, text="Room Management", font=("Arial", 24)).pack(pady=20)
        
        input_frame = ctk.CTkFrame(frame)
        input_frame.pack(pady=10)
        
        name_ent = ctk.CTkEntry(input_frame, placeholder_text="Room No (e.g. VIB 503)")
        name_ent.pack(side="left", padx=5)
        cap_ent = ctk.CTkEntry(input_frame, placeholder_text="Capacity")
        cap_ent.pack(side="left", padx=5)
        type_ent = ctk.CTkComboBox(input_frame, values=["Lecture Hall", "Lab"])
        type_ent.pack(side="left", padx=5)

        def add_rm():
            if name_ent.get() and cap_ent.get().isdigit():
                self.db.add_room(name_ent.get(), int(cap_ent.get()), type_ent.get())
                self.refresh_list(list_box, "rooms")
                name_ent.delete(0, 'end')
                cap_ent.delete(0, 'end')
            else:
                messagebox.showerror("Error", "Invalid Input")

        ctk.CTkButton(input_frame, text="Add", command=add_rm).pack(side="left", padx=5)
        list_box = ctk.CTkTextbox(frame, width=700, height=400)
        list_box.pack(pady=20)
        self.refresh_list(list_box, "rooms")
        return frame

    def create_settings_frame(self):
        frame = ctk.CTkFrame(self)
        ctk.CTkLabel(frame, text="Application Settings", font=("Arial", 24)).pack(pady=20)
        
        # Scaling
        ctk.CTkLabel(frame, text="UI Scaling:").pack(pady=(20,5))
        scale_opt = ctk.CTkOptionMenu(frame, values=["80%", "90%", "100%", "110%", "120%"], 
                                      command=self.change_scaling)
        scale_opt.set("100%")
        scale_opt.pack(pady=5)
        
        # Appearance
        ctk.CTkLabel(frame, text="Appearance Mode:").pack(pady=(20,5))
        app_opt = ctk.CTkOptionMenu(frame, values=["System", "Light", "Dark"], 
                                    command=lambda v: ctk.set_appearance_mode(v))
        app_opt.set("System")
        app_opt.pack(pady=5)
        
        return frame

    # --- ACTIONS ---
    def add_teacher_action(self, entries, list_box):
        name = entries[0].get()
        code = entries[1].get()
        if name and code:
            if self.db.add_teacher(name, code):
                entries[0].delete(0, 'end')
                entries[1].delete(0, 'end')
                self.refresh_list(list_box, "teachers")
            else:
                messagebox.showerror("Error", "Teacher Code must be unique.")
        else:
            messagebox.showerror("Error", "All fields are required.")
            
    def add_section_action(self, entries, list_box):
        name = entries[0].get()
        if name:
            if self.db.add_section(name):
                entries[0].delete(0, 'end')
                self.refresh_list(list_box, "sections")
            else:
                messagebox.showerror("Error", "Section Name must be unique.")

    def refresh_list(self, text_widget, table):
        data = self.db.fetch_all(table)
        text_widget.configure(state="normal")
        text_widget.delete("1.0", "end")
        
        headers = {
            "teachers": "ID | Name | Code",
            "subjects": "ID | Name | Code | Type | Hours",
            "rooms": "ID | Name | Capacity | Type",
            "sections": "ID | Name"
        }
        
        text_widget.insert("end", headers.get(table, "") + "\n" + "-"*50 + "\n")
        
        for row in data:
            row_str = " | ".join(str(item) for item in row)
            text_widget.insert("end", row_str + "\n")
            
        text_widget.configure(state="disabled")

    def show_frame(self, name):
        # Update Dashboard Stats on switch
        if name == "Dashboard":
            t = len(self.db.fetch_all("teachers"))
            s = len(self.db.fetch_all("subjects"))
            r = len(self.db.fetch_all("rooms"))
            sec = len(self.db.fetch_all("sections"))
            self.stats_label.configure(text=f"Database Status:\nTeachers: {t}\nSubjects: {s}\nRooms: {r}\nSections: {sec}")

        # Hide all, show selected
        for frame in self.frames.values():
            frame.grid_forget()
        self.frames[name].grid(row=0, column=1, sticky="nsew", padx=20, pady=20)

    def change_scaling(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        ctk.set_widget_scaling(new_scaling_float)

    def run_generation(self):
        self.preview_area.configure(state="normal")
        self.preview_area.delete("1.0", "end")
        self.preview_area.insert("end", "Generating schedule... please wait...\n")
        self.update()
        
        schedule, status = self.scheduler.generate()
        
        if schedule:
            self.schedule_data = schedule
            self.preview_area.insert("end", "Generation Successful!\n\n")
            self.preview_area.insert("end", "-"*30 + "\nPREVIEW (Sample Data)\n" + "-"*30 + "\n")
            
            # Show a snippet in preview
            days = ["Monday", "Tuesday"]
            for day in days:
                self.preview_area.insert("end", f"\n[ {day} ]\n")
                for slot in self.scheduler.slots:
                    if schedule[day][slot]:
                        first_sec_id = list(schedule[day][slot].keys())[0]
                        info = schedule[day][slot][first_sec_id]
                        self.preview_area.insert("end", f"{slot}: {info['subject']} by {info['teacher']} in {info['room']}\n")
                    else:
                        self.preview_area.insert("end", f"{slot}: FREE\n")
            
            self.export_btn.configure(state="normal")
        else:
            self.preview_area.insert("end", f"Failed: {status}")
            self.export_btn.configure(state="disabled")
            
        self.preview_area.configure(state="disabled")

    def export_file(self):
        if not self.schedule_data: return
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                filetypes=[("Excel file", "*.xlsx")])
        if filename:
            try:
                self.scheduler.export_to_excel(self.schedule_data, filename)
                messagebox.showinfo("Success", "Timetable exported successfully!")
            except Exception as e:
                messagebox.showerror("Error", f"Could not save file: {e}")

if __name__ == "__main__":
    app = TimeTableApp()
    app.mainloop()
